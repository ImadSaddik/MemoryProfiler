import openpyxl

from memory_profiler import profile


@profile
def extract_content_from_excel_file(file_path: str) -> str:
    workbook = openpyxl.load_workbook(
        file_path, read_only=True, data_only=True)
    file_content = ""
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            row_values = []
            for cell in row:
                cell_value = str(cell.value).strip(
                ) if cell.value is not None else ''
                cleaned_value = " ".join(cell_value.split())

                if cleaned_value:
                    row_values.append(cleaned_value)

            row_text = " ".join(row_values)
            if row_text:
                file_content += row_text + "\n"

    return file_content.strip()


if __name__ == "__main__":
    file_path = "./data.xlsx"
    content = extract_content_from_excel_file(file_path)
    print(len(content))
